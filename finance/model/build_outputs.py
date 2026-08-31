"""
Build the Mulligan Mints financial outputs.

    python3 build_outputs.py

Writes:
    ../outputs/mulligan-mints-5yr-model.xlsx
    ../outputs/model-summary.md
"""
from __future__ import annotations

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import assumptions as A
import model as M

OUT = os.path.join(os.path.dirname(__file__), "..", "outputs")
os.makedirs(OUT, exist_ok=True)

GREEN = "0B3D2E"
ACCENT = "C8A951"
LIGHT = "EAF0EC"

H1 = Font(bold=True, size=14, color="FFFFFF")
H2 = Font(bold=True, size=11, color="FFFFFF")
BOLD = Font(bold=True)
FILL_H = PatternFill("solid", fgColor=GREEN)
FILL_SUB = PatternFill("solid", fgColor=LIGHT)
THIN = Side(style="thin", color="BBBBBB")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

MONEY = '#,##0;[Red](#,##0)'
MONEY2 = '#,##0.00;[Red](#,##0.00)'
PCT = '0.0%'
UNITS = '#,##0'


def sheet_title(ws, text, ncols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = H1
    c.fill = FILL_H
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def header_row(ws, row, values, start_col=1):
    for j, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + j, value=v)
        c.font = H2
        c.fill = FILL_H
        c.alignment = Alignment(horizontal="center" if j else "left")
        c.border = BOX


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# unit economics
# ---------------------------------------------------------------------------

def unit_economics(scenario="base", year=1):
    """Per-tin waterfall for a representative tin in a given model year."""
    r = M.run(scenario)
    sc = A.SCENARIOS[scenario]
    m = (year - 1) * 12 + 6
    rate = M.usdzar(m)
    annual = r.annual("units_total")[year - 1]
    fob_usd = M.fob_usd_for_annual_volume(annual) * sc["fob_multiplier"]

    fob = fob_usd * rate
    freight = A.FREIGHT_USD_PER_20FT / A.UNITS_PER_20FT * rate
    ins = fob * A.MARINE_INSURANCE_PCT_OF_FOB
    duty = fob * A.IMPORT_DUTY_PCT
    clearing = A.CLEARING_ZAR_PER_20FT / A.UNITS_PER_20FT
    landed = fob + freight + ins + duty + clearing
    landed_w = landed / (1 - A.STOCK_WASTAGE_PCT)

    pf = M.annual_price_factor(m) * sc["net_price_multiplier"]
    rows = []
    for k in ("golf", "bars", "dtc", "retail", "export"):
        net = A.CHANNELS[k]["net_price_y1"] * pf
        logi = (A.LOGISTICS_ZAR_PER_UNIT_EXPORT if k == "export" else A.LOGISTICS_ZAR_PER_UNIT)
        logi = M.inflate(logi, A.OPEX_INFLATION_PA, m)
        comm = net * A.SALES_COMMISSION_PCT if k in ("golf", "bars") else 0.0
        contrib = net - landed_w - logi - comm
        rows.append({
            "channel": A.CHANNELS[k]["label"],
            "net_price": net,
            "landed_cost": landed_w,
            "gross_profit": net - landed_w,
            "gross_margin": (net - landed_w) / net,
            "logistics": logi,
            "commission": comm,
            "contribution": contrib,
            "contribution_margin": contrib / net,
        })
    stack = {
        "USD/ZAR": rate, "FOB USD/tin": fob_usd, "FOB ZAR": fob,
        "Sea freight": freight, "Marine insurance": ins,
        "Import duty (25% of FOB)": duty, "Clearing & inland": clearing,
        "Landed cost": landed, "Landed + wastage allowance": landed_w,
    }
    return stack, rows


# ---------------------------------------------------------------------------
# workbook
# ---------------------------------------------------------------------------

def build_workbook():
    wb = Workbook()
    results = {s: M.run(s) for s in ("bootstrap", "bear", "base", "bull")}
    needs = {s: M.funding_need(s) for s in ("bootstrap", "bear", "base", "bull")}
    base = results["base"]

    # ---------------- cover ----------------
    ws = wb.active
    ws.title = "Read me"
    sheet_title(ws, "MULLIGAN MINTS  —  5-YEAR FINANCIAL MODEL", 6)
    widths(ws, {"A": 34, "B": 74})
    lines = [
        ("Built", date.today().isoformat()),
        ("Base currency", "ZAR, excluding VAT unless stated"),
        ("Horizon", f"{A.HORIZON_MONTHS} months from funding close ({A.MONTH_ONE_LABEL})"),
        ("Source of truth", "finance/model/assumptions.py — change numbers there, not in this workbook"),
        ("Rebuild", "cd finance/model && python3 build_outputs.py"),
        ("", ""),
        ("HEALTH WARNING", "Every cost marked [EST] in assumptions.py is an estimate. The two that "
                           "move the answer most are the supplier FOB price and the sell-through rate "
                           "per venue per month. Replace both with real numbers as soon as the Suntak "
                           "quote and the first 90 days of venue data land."),
        ("", ""),
        ("Scenarios", "bear / base / bull — see the Scenarios sheet"),
    ]
    r = 3
    for k, v in lines:
        ws.cell(row=r, column=1, value=k).font = BOLD
        c = ws.cell(row=r, column=2, value=v)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if len(v) > 90:
            ws.row_dimensions[r].height = 58
        r += 1

    # ---------------- assumptions ----------------
    ws = wb.create_sheet("Assumptions")
    sheet_title(ws, "KEY ASSUMPTIONS", 5)
    widths(ws, {"A": 46, "B": 16, "C": 62})
    r = 3
    blocks = [
        ("Macro", [
            ("USD/ZAR year 1", A.USDZAR_YEAR_1, "Spot ~15.93 (Aug 2026); planned at 16.00"),
            ("ZAR depreciation p.a.", A.ZAR_DEPRECIATION_PA, "Long-run drift — understating this is the importer's classic error"),
            ("Opex inflation p.a.", A.OPEX_INFLATION_PA, "Wages and logistics run above CPI"),
            ("Price increase p.a.", A.PRICE_INCREASE_PA, "Annual list price increase"),
            ("Company tax rate", A.CORPORATE_TAX_RATE, "SARS; assessed losses offset max 80% of taxable income"),
        ]),
        ("Product & landed cost", [
            ("Tins per carton", A.UNITS_PER_CARTON, "12 per display box x 8 boxes (supplier spec)"),
            ("Tins per 20ft container", A.UNITS_PER_20FT, "1,400 cartons per 20ft (supplier spec)"),
            ("FOB USD/tin at launch volume", A.FOB_USD_CURVE[0][1], "[EST] REPLACE with the Suntak quote"),
            ("FOB USD/tin at scale", A.FOB_USD_CURVE[-1][1], "[EST] volume curve in assumptions.py"),
            ("Sea freight USD per 20ft", A.FREIGHT_USD_PER_20FT, "$2,403-$2,937 China->Durban/CT, Jul 2026"),
            ("Import duty", A.IMPORT_DUTY_PCT, "HS 1704 sugar confectionery, on FOB customs value"),
            ("Clearing ZAR per 20ft", A.CLEARING_ZAR_PER_20FT, "[EST] port, agency, unpack, inland"),
            ("Stock wastage allowance", A.STOCK_WASTAGE_PCT, "[EST] damage, shorts, expiry"),
        ]),
        ("Route to market", [
            ("Target RSP incl VAT", A.TARGET_RSP_INCL_VAT, "vs Fisherman's Friend 25g at R28.99 (Dis-Chem)"),
            ("Net price — golf & bars direct", A.CHANNELS["golf"]["net_price_y1"], "Venue takes ~38% of ex-VAT RSP"),
            ("Net price — distributor (bars)", A.DISTRIBUTOR_NET_PRICE_BARS, "Reach bought with ~R6/tin of margin"),
            ("Net price — grocery/pharmacy", A.CHANNELS["retail"]["net_price_y1"], "After 32% retail margin + 13% trade spend"),
            ("Net price — export FOB SA", A.CHANNELS["export"]["net_price_y1"], "Importer carries freight, duty, local marketing"),
            ("Golf club universe", A.CHANNELS["golf"]["universe"], "460 GolfRSA affiliated clubs; ~300 serviceable"),
            ("Bar/restaurant universe", A.CHANNELS["bars"]["universe"], "5,178 formal F&B enterprises (Stats SA); metro premium subset"),
        ]),
        ("Working capital", [
            ("Supplier deposit", A.SUPPLIER_DEPOSIT_PCT, "Balance against bill of lading"),
            ("Order-to-shelf lead time (months)", M.LEAD_MONTHS, "Production 2 + transit 1 + clearing 0.5"),
            ("Target forward cover (months)", A.TARGET_FORWARD_COVER_MONTHS, "Lead time plus safety stock"),
            ("Minimum order (units)", A.MIN_ORDER_UNITS, "[EST] 3 flavours x per-SKU MOQ"),
        ]),
    ]
    for title, items in blocks:
        c = ws.cell(row=r, column=1, value=title)
        c.font = BOLD
        for col in range(1, 4):
            ws.cell(row=r, column=col).fill = FILL_SUB
        r += 1
        for label, val, note in items:
            ws.cell(row=r, column=1, value=label)
            vc = ws.cell(row=r, column=2, value=val)
            vc.number_format = PCT if isinstance(val, float) and val < 1 else MONEY2
            ws.cell(row=r, column=3, value=note).alignment = Alignment(wrap_text=False)
            r += 1
        r += 1

    # ---------------- unit economics ----------------
    ws = wb.create_sheet("Unit economics")
    sheet_title(ws, "UNIT ECONOMICS PER 35g TIN  (base case, year 1)", 9)
    widths(ws, {"A": 44, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14})
    stack, rows = unit_economics("base", 1)
    r = 3
    ws.cell(row=r, column=1, value="LANDED COST BUILD-UP (ZAR per tin)").font = BOLD
    r += 1
    for k, v in stack.items():
        ws.cell(row=r, column=1, value=k)
        c = ws.cell(row=r, column=2, value=round(v, 4))
        c.number_format = MONEY2
        if k.startswith("Landed"):
            ws.cell(row=r, column=1).font = BOLD
            c.font = BOLD
        r += 1

    r += 1
    header_row(ws, r, ["Channel", "Net price", "Landed cost", "Gross profit",
                       "Gross margin", "Logistics", "Commission",
                       "Contribution", "Contribution %"])
    r += 1
    for row in rows:
        ws.cell(row=r, column=1, value=row["channel"])
        for j, key in enumerate(["net_price", "landed_cost", "gross_profit", "gross_margin",
                                 "logistics", "commission", "contribution", "contribution_margin"], start=2):
            c = ws.cell(row=r, column=j, value=round(row[key], 4))
            c.number_format = PCT if "margin" in key else MONEY2
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="THE NUMBERS THAT MATTER").font = BOLD
    r += 1
    b = M.run("base")
    fixed_m12 = b.rows["opex_salaries"][11] + b.rows["opex_overheads"][11]
    mktg_m12 = b.rows["opex_marketing"][11]
    contrib = rows[0]["contribution"]
    be_all = (fixed_m12 + mktg_m12) / contrib
    be_fixed = fixed_m12 / contrib
    for note in [
        f"A golf club buying a 12-tin display box pays ~R{rows[0]['net_price']*12:,.0f} ex-VAT "
        f"and sells it for ~R{A.TARGET_RSP_INCL_VAT*12:,.0f} incl VAT.",
        f"Contribution per direct tin: R{contrib:,.2f}.",
        f"At the month-12 cost base, break-even is {be_all:,.0f} tins/month including marketing "
        f"({be_fixed:,.0f} excluding it).",
        f"At {A.CHANNELS['golf']['units_per_outlet_month'][1]} tins per outlet per month that is "
        f"~{be_all/A.CHANNELS['golf']['units_per_outlet_month'][1]:,.0f} active outlets. "
        "The year-1 plan exits with ~250. Break-even therefore lands early in year 2, not in year 1 —",
        "and the model agrees: first EBITDA-positive month is month 21.",
        "This is the single most important sanity check in the plan. If venues sell through at half",
        "this rate, break-even needs twice the outlets, and no amount of marketing spend fixes it.",
    ]:
        ws.cell(row=r, column=1, value=note)
        r += 1

    # ---------------- annual P&L ----------------
    ws = wb.create_sheet("P&L annual")
    sheet_title(ws, "PROFIT & LOSS — ANNUAL, ALL SCENARIOS (ZAR)", 7)
    widths(ws, {"A": 42, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
    r = 3
    pl_lines = [
        ("Units sold", "units_total", UNITS),
        ("", None, None),
        ("Revenue — golf clubs", "rev_golf", MONEY),
        ("Revenue — bars & restaurants", "rev_bars", MONEY),
        ("Revenue — direct to consumer", "rev_dtc", MONEY),
        ("Revenue — grocery & pharmacy retail", "rev_retail", MONEY),
        ("Revenue — export", "rev_export", MONEY),
        ("Revenue — brand licensing", "rev_licensing", MONEY),
        ("NET REVENUE", "revenue", MONEY),
        ("Cost of goods sold", "cogs", MONEY),
        ("GROSS PROFIT", "gross_profit", MONEY),
        ("", None, None),
        ("Salaries & employer costs", "opex_salaries", MONEY),
        ("Marketing, trade & listing fees", "opex_marketing", MONEY),
        ("Logistics & warehousing", "opex_logistics", MONEY),
        ("Sales commission", "opex_commission", MONEY),
        ("Product & export development", "opex_npd", MONEY),
        ("Fixed overheads", "opex_overheads", MONEY),
        ("One-off setup costs", "opex_setup", MONEY),
        ("TOTAL OPERATING COSTS", "opex_total", MONEY),
        ("", None, None),
        ("EBITDA", "ebitda", MONEY),
        ("Finance costs (trade finance interest)", "finance_cost", MONEY),
        ("Tax charge", "tax_charge", MONEY),
        ("NET PROFIT", "net_profit", MONEY),
    ]
    for scen in ("base", "bootstrap", "bear", "bull"):
        res = results[scen]
        c = ws.cell(row=r, column=1, value=f"{scen.upper()} — {A.SCENARIOS[scen]['label']}")
        c.font = BOLD
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = FILL_SUB
        r += 1
        header_row(ws, r, ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
        r += 1
        for label, key, fmt in pl_lines:
            if key is None:
                r += 1
                continue
            lc = ws.cell(row=r, column=1, value=label)
            if label.isupper():
                lc.font = BOLD
            for y, v in enumerate(res.annual(key), start=2):
                cell = ws.cell(row=r, column=y, value=round(v))
                cell.number_format = fmt
                if label.isupper():
                    cell.font = BOLD
            r += 1
        # margins
        for label, num, den in [("Gross margin %", "gross_profit", "revenue"),
                                ("EBITDA margin %", "ebitda", "revenue")]:
            ws.cell(row=r, column=1, value=label).font = BOLD
            for y, (n, d) in enumerate(zip(res.annual(num), res.annual(den)), start=2):
                cell = ws.cell(row=r, column=y, value=(n / d if d else 0))
                cell.number_format = PCT
                cell.font = BOLD
            r += 1
        r += 2

    # ---------------- monthly P&L and cashflow ----------------
    for title, keys, name in [
        ("MONTHLY PROFIT & LOSS — BASE CASE (ZAR)", [
            ("Units sold", "units_total", UNITS),
            ("Net revenue", "revenue", MONEY),
            ("Cost of goods sold", "cogs", MONEY),
            ("Gross profit", "gross_profit", MONEY),
            ("Salaries", "opex_salaries", MONEY),
            ("Marketing & listings", "opex_marketing", MONEY),
            ("Logistics", "opex_logistics", MONEY),
            ("Commission", "opex_commission", MONEY),
            ("Product & export dev", "opex_npd", MONEY),
            ("Overheads", "opex_overheads", MONEY),
            ("Setup costs", "opex_setup", MONEY),
            ("Total operating costs", "opex_total", MONEY),
            ("EBITDA", "ebitda", MONEY),
            ("Headcount", "headcount", UNITS),
        ], "P&L monthly"),
        ("MONTHLY CASHFLOW — BASE CASE (ZAR)", [
            ("Receipts from customers (incl VAT)", "cash_in_sales", MONEY),
            ("Payments to supplier (FOB, freight, insurance)", "cash_out_supplier", MONEY),
            ("Import duty & clearing", "cash_out_duty_clearing", MONEY),
            ("Operating costs paid", "cash_out_opex", MONEY),
            ("VAT settled", "cash_vat", MONEY),
            ("Tax paid", "cash_tax", MONEY),
            ("Trade finance drawn", "cash_tf_draw", MONEY),
            ("Trade finance repaid", "cash_tf_repay", MONEY),
            ("Investor revenue share", "cash_investor_repay", MONEY),
            ("Dividends declared", "cash_dividend", MONEY),
            ("Funding received", "cash_funding", MONEY),
            ("NET CASHFLOW", "net_cashflow", MONEY),
            ("CLOSING CASH", "closing_cash", MONEY),
            ("Stock on hand (units)", "stock_units", UNITS),
            ("Stock on hand (value)", "stock_value", MONEY),
            ("Debtors", "debtors", MONEY),
            ("Trade finance outstanding", "tf_outstanding", MONEY),
            ("Investor capital outstanding", "investor_outstanding", MONEY),
            ("  of dividends, to investor", "dividend_to_investor", MONEY),
        ], "Cashflow monthly"),
    ]:
        ws = wb.create_sheet(name)
        sheet_title(ws, title, 61)
        ws.column_dimensions["A"].width = 40
        ws.freeze_panes = "B4"
        header_row(ws, 3, ["Month"] + [f"M{m}" for m in base.months])
        for j in range(2, 62):
            ws.column_dimensions[get_column_letter(j)].width = 12
        r = 4
        for label, key, fmt in keys:
            lc = ws.cell(row=r, column=1, value=label)
            if label.isupper():
                lc.font = BOLD
            for j, v in enumerate(base.rows[key], start=2):
                c = ws.cell(row=r, column=j, value=round(v))
                c.number_format = fmt
                if label.isupper():
                    c.font = BOLD
            r += 1

    # ---------------- scenarios ----------------
    ws = wb.create_sheet("Scenarios")
    sheet_title(ws, "SCENARIO COMPARISON", 5)
    widths(ws, {"A": 40, "B": 18, "C": 18, "D": 18, "E": 18})
    r = 3
    header_row(ws, r, ["", "BOOTSTRAP", "BEAR", "BASE", "BULL"])
    r += 1
    comps = [
        ("Year 5 units vs base", lambda s: (results[s].annual("units_total")[4]
                                            / results["base"].annual("units_total")[4]), PCT),
        ("Year 1 units", lambda s: results[s].annual("units_total")[0], UNITS),
        ("Year 3 units", lambda s: results[s].annual("units_total")[2], UNITS),
        ("Year 5 units", lambda s: results[s].annual("units_total")[4], UNITS),
        ("Year 1 revenue", lambda s: results[s].annual("revenue")[0], MONEY),
        ("Year 3 revenue", lambda s: results[s].annual("revenue")[2], MONEY),
        ("Year 5 revenue", lambda s: results[s].annual("revenue")[4], MONEY),
        ("Year 5 gross margin", lambda s: results[s].annual("gross_profit")[4] / results[s].annual("revenue")[4], PCT),
        ("Year 5 EBITDA", lambda s: results[s].annual("ebitda")[4], MONEY),
        ("Year 5 EBITDA margin", lambda s: results[s].annual("ebitda")[4] / results[s].annual("revenue")[4], PCT),
        ("Cumulative 5-year EBITDA", lambda s: sum(results[s].annual("ebitda")), MONEY),
        ("First EBITDA-positive year", lambda s: next(
            (i + 1 for i, v in enumerate(results[s].annual("ebitda")) if v > 0), 0),
         '0;;"never"'),
        ("Unfunded peak cash deficit", lambda s: needs[s]["peak_deficit"], MONEY),
        ("Month of peak deficit", lambda s: needs[s]["peak_month"], '0'),
        ("Lowest funded cash balance", lambda s: needs[s]["min_closing_cash"], MONEY),
        ("Year 5 closing cash", lambda s: results[s].annual_last("closing_cash")[4], MONEY),
    ]
    for label, fn, fmt in comps:
        ws.cell(row=r, column=1, value=label).font = BOLD
        for j, s in enumerate(("bootstrap", "bear", "base", "bull"), start=2):
            c = ws.cell(row=r, column=j, value=round(fn(s), 4))
            c.number_format = fmt
        r += 1

    # ---------------- funding ----------------
    ws = wb.create_sheet("Funding")
    sheet_title(ws, "FUNDING & USE OF PROCEEDS", 5)
    widths(ws, {"A": 52, "B": 18, "C": 62})
    r = 3
    ws.cell(row=r, column=1, value="ROUNDS").font = BOLD
    r += 1
    header_row(ws, r, ["Round", "Amount (ZAR)", "Month"])
    r += 1
    for m, label, amt in A.FUNDING_ROUNDS:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=amt).number_format = MONEY
        ws.cell(row=r, column=3, value=f"Month {m}")
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="USE OF THE PRE-SEED (first 12 months, base case)").font = BOLD
    r += 1
    header_row(ws, r, ["Line", "ZAR", "Note"])
    r += 1
    setup_total = sum(c for _m, _i, c in A.SETUP_COSTS)
    y1 = base
    first_po = base.purchase_orders[0] if base.purchase_orders else None
    use = [
        ("Brand, trademark, artwork, content, website, launch", setup_total,
         "Itemised in assumptions.SETUP_COSTS — PJ Offner identity is R150k of this"),
        ("First production order (FOB, freight, insurance)",
         (first_po.fob_zar_total + first_po.freight_zar + first_po.insurance_zar) if first_po else 0,
         f"{first_po.units:,} tins" if first_po else ""),
        ("Import duty & clearing on first order",
         (first_po.duty_zar + first_po.clearing_zar) if first_po else 0,
         "25% duty on FOB customs value plus port costs"),
        ("Salaries, months 1-12", sum(y1.rows["opex_salaries"][:12]), "Founder from month 4, first rep from month 7"),
        ("Marketing, months 1-12", sum(y1.rows["opex_marketing"][:12]), "Creator seeding, club days, sampling, POS"),
        ("Overheads & logistics, months 1-12",
         sum(y1.rows["opex_overheads"][:12]) + sum(y1.rows["opex_logistics"][:12]), ""),
    ]
    total_use = 0
    for label, amt, note in use:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=round(amt)).number_format = MONEY
        ws.cell(row=r, column=3, value=note)
        total_use += amt
        r += 1
    ws.cell(row=r, column=1, value="Sub-total, year 1 committed").font = BOLD
    c = ws.cell(row=r, column=2, value=round(total_use)); c.number_format = MONEY; c.font = BOLD
    r += 1
    ws.cell(row=r, column=1, value="Second production order + working capital buffer").font = BOLD
    c = ws.cell(row=r, column=2, value=round(A.FUNDING_ROUNDS[0][2] - total_use)); c.number_format = MONEY; c.font = BOLD
    r += 2
    ws.cell(row=r, column=1, value="Note: the base case dips to its lowest cash balance in month "
            f"{needs['base']['min_closing_month']} at R{needs['base']['min_closing_cash']:,.0f}. "
            "The growth round must be closed before then, not started then.")

    # ---------------- purchase orders ----------------
    ws = wb.create_sheet("Purchase orders")
    sheet_title(ws, "PURCHASE ORDER SCHEDULE — BASE CASE", 10)
    widths(ws, {c: 16 for c in "ABCDEFGHIJ"})
    header_row(ws, 3, ["PO month", "Arrives", "Units", "20ft equiv", "FOB ZAR",
                       "Freight", "Insurance", "Duty", "Clearing", "Landed/tin"])
    r = 4
    for po in base.purchase_orders:
        vals = [po.order_month, po.arrive_month, po.units,
                round(po.units / A.UNITS_PER_20FT, 2), round(po.fob_zar_total),
                round(po.freight_zar), round(po.insurance_zar), round(po.duty_zar),
                round(po.clearing_zar), round(po.landed_per_unit, 2)]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.number_format = MONEY2 if j == 10 else (UNITS if j in (1, 2, 3) else MONEY)
        r += 1

    # ---------------- exit ----------------
    try:
        import exit_analysis as EX
        ws = wb.create_sheet("Exit")
        sheet_title(ws, "POTENTIAL EXIT AT MONTH 60 — BOOTSTRAP SCENARIO", 8)
        widths(ws, {"A": 40, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
        bs = results["bootstrap"]
        r_ = 3
        ws.cell(row=r_, column=1, value="WHAT IS BEING SOLD").font = BOLD
        r_ += 1
        w0 = EX.waterfall(EX.EBITDA_MULTIPLES[0])
        for label, val in [
            ("Year-5 EBITDA", bs.annual("ebitda")[4]),
            ("Year-5 revenue", bs.annual("revenue")[4]),
            ("  of which brand licensing", bs.annual("rev_licensing")[4]),
            ("Cash at bank", bs.annual_last("closing_cash")[4]),
            (f"  less working capital left in ({EX.WORKING_CAPITAL_PEG_MONTHS} months opex)", -w0["wc_peg"]),
            ("  surplus cash to sellers", w0["surplus_cash"]),
            ("Trade finance owed", -w0["trade_finance"]),
            ("Investor capital still owed", -w0["investor_residual"]),
        ]:
            ws.cell(row=r_, column=1, value=label)
            ws.cell(row=r_, column=2, value=round(val)).number_format = MONEY
            r_ += 1
        r_ += 1
        header_row(ws, r_, ["EBITDA multiple", "Enterprise value", "Equity value",
                            "Founders 80%", "Founders after CGT", "PJ Offner 10%",
                            "Investor at exit", "Investor IRR"])
        r_ += 1
        for mult in EX.EBITDA_MULTIPLES:
            w = EX.waterfall(mult)
            vals = [f"{mult}x", w["enterprise_value"], w["equity_value"], w["founders_gross"],
                    w["founders_net"], w["pj_gross"], w["investor_total_at_exit"],
                    EX.investor_irr(mult)]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r_, column=j, value=v if j == 1 else round(v, 4))
                c.number_format = PCT if j == 8 else (MONEY if j > 1 else '@')
            r_ += 1
        r_ += 2
        for note in [
            "Equity value = enterprise value + surplus cash - trade finance - investor capital owed.",
            "Founders' CGT at 18% effective (40% inclusion x 45% marginal, individuals holding shares).",
            "The investor is also repaid R{:,.0f} of capital at R1 a tin across the five years,".format(
                w0["investor_repaid_over_plan"]),
            "so their all-in return runs from R{:,.0f} at 4x to R{:,.0f} at 10x, on R1,000,000.".format(
                EX.waterfall(4)["investor_all_in"], EX.waterfall(10)["investor_all_in"]),
            "",
            "Every multiple here is an estimate. Nobody has offered anything.",
        ]:
            ws.cell(row=r_, column=1, value=note)
            r_ += 1
    except Exception as exc:      # never let the exit sheet break the build
        print("exit sheet skipped:", exc)

    path = os.path.join(OUT, "mulligan-mints-5yr-model.xlsx")
    wb.save(path)
    return path, results, needs


# ---------------------------------------------------------------------------
# markdown summary
# ---------------------------------------------------------------------------

def rands(v):
    if abs(v) >= 1_000_000:
        return f"R{v/1_000_000:,.2f}m"
    return f"R{v:,.0f}"


def build_markdown(results, needs):
    base = results["base"]
    stack, ue = unit_economics("base", 1)
    L = []
    w = L.append

    w("# Mulligan Mints — model output summary\n")
    w("> Generated by `finance/model/build_outputs.py`. **Do not edit by hand** — "
      "change `finance/model/assumptions.py` and rebuild.\n")
    w(f"Built {date.today().isoformat()}. All figures ZAR, excluding VAT unless stated.\n")

    w("## 1. Unit economics — one 35g tin\n")
    w("| Landed cost build-up | ZAR/tin |")
    w("| --- | ---: |")
    for k, v in stack.items():
        w(f"| {k} | {v:,.2f} |")
    w("")
    w("| Channel | Net price | Landed cost | Gross profit | Gross margin | Contribution | Contribution % |")
    w("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for row in ue:
        w(f"| {row['channel']} | {row['net_price']:.2f} | {row['landed_cost']:.2f} | "
          f"{row['gross_profit']:.2f} | {row['gross_margin']*100:.0f}% | "
          f"{row['contribution']:.2f} | {row['contribution_margin']*100:.0f}% |")
    w("")

    w("## 2. Base case — five-year P&L\n")
    w("| ZAR | Year 1 | Year 2 | Year 3 | Year 4 | Year 5 |")
    w("| --- | ---: | ---: | ---: | ---: | ---: |")
    def line(label, key, fmt=rands):
        vals = base.annual(key)
        w(f"| {label} | " + " | ".join(fmt(v) for v in vals) + " |")
    w("| **Units sold** | " + " | ".join(f"{v:,.0f}" for v in base.annual("units_total")) + " |")
    for label, key in [("Revenue — golf clubs", "rev_golf"),
                       ("Revenue — bars & restaurants", "rev_bars"),
                       ("Revenue — direct to consumer", "rev_dtc"),
                       ("Revenue — grocery & pharmacy", "rev_retail"),
                       ("Revenue — export", "rev_export"),
                       ("Revenue — brand licensing", "rev_licensing")]:
        line(label, key)
    line("**Net revenue**", "revenue")
    line("Cost of goods sold", "cogs")
    line("**Gross profit**", "gross_profit")
    w("| Gross margin | " + " | ".join(
        f"{(g/r*100 if r else 0):.0f}%" for g, r in zip(base.annual("gross_profit"), base.annual("revenue"))) + " |")
    for label, key in [("Salaries", "opex_salaries"),
                       ("Marketing, trade & listings", "opex_marketing"),
                       ("Logistics & warehousing", "opex_logistics"),
                       ("Sales commission", "opex_commission"),
                       ("Product & export development", "opex_npd"),
                       ("Fixed overheads", "opex_overheads"),
                       ("One-off setup", "opex_setup")]:
        line(label, key)
    line("**Total operating costs**", "opex_total")
    line("**EBITDA**", "ebitda")
    line("Finance costs (trade finance)", "finance_cost")
    w("| EBITDA margin | " + " | ".join(
        f"{(e/r*100 if r else 0):.0f}%" for e, r in zip(base.annual("ebitda"), base.annual("revenue"))) + " |")
    line("Tax charge", "tax_charge")
    line("**Net profit**", "net_profit")
    w("")

    w("## 3. Base case — cashflow\n")
    w("| ZAR | Year 1 | Year 2 | Year 3 | Year 4 | Year 5 |")
    w("| --- | ---: | ---: | ---: | ---: | ---: |")
    for label, key in [("Receipts from customers (incl VAT)", "cash_in_sales"),
                       ("Payments to supplier", "cash_out_supplier"),
                       ("Import duty & clearing", "cash_out_duty_clearing"),
                       ("Operating costs paid", "cash_out_opex"),
                       ("VAT settled", "cash_vat"),
                       ("Tax paid", "cash_tax"),
                       ("Trade finance drawn", "cash_tf_draw"),
                       ("Trade finance repaid", "cash_tf_repay"),
                       ("Investor revenue share", "cash_investor_repay"),
                       ("Dividends declared", "cash_dividend"),
                       ("Funding received", "cash_funding"),
                       ("**Net cashflow**", "net_cashflow")]:
        line(label, key)
    w("| **Closing cash** | " + " | ".join(rands(v) for v in base.annual_last("closing_cash")) + " |")
    w("| Stock on hand (units) | " + " | ".join(f"{v:,.0f}" for v in base.annual_last("stock_units")) + " |")
    w("| Stock on hand (value) | " + " | ".join(rands(v) for v in base.annual_last("stock_value")) + " |")
    w("")

    w("## 4. Scenarios\n")
    w("| | Bootstrap | Bear | Base | Bull |")
    w("| --- | ---: | ---: | ---: | ---: |")
    rows = [
        ("Year 5 units vs base", lambda s: f"{results[s].annual('units_total')[4]/results['base'].annual('units_total')[4]:.0%}"),
        ("Year 1 units", lambda s: f"{results[s].annual('units_total')[0]:,.0f}"),
        ("Year 5 units", lambda s: f"{results[s].annual('units_total')[4]:,.0f}"),
        ("Year 1 revenue", lambda s: rands(results[s].annual("revenue")[0])),
        ("Year 3 revenue", lambda s: rands(results[s].annual("revenue")[2])),
        ("Year 5 revenue", lambda s: rands(results[s].annual("revenue")[4])),
        ("Year 5 EBITDA", lambda s: rands(results[s].annual("ebitda")[4])),
        ("Year 5 EBITDA margin", lambda s: f"{results[s].annual('ebitda')[4]/results[s].annual('revenue')[4]*100:.0f}%"),
        ("Cumulative 5-yr EBITDA", lambda s: rands(sum(results[s].annual("ebitda")))),
        ("First EBITDA-positive year", lambda s: (
            lambda y: str(y) if y else "never")(
            next((i+1 for i, v in enumerate(results[s].annual("ebitda")) if v > 0), 0))),
        ("Unfunded peak cash deficit", lambda s: rands(needs[s]["peak_deficit"])),
        ("Month of peak deficit", lambda s: f"M{needs[s]['peak_month']}"),
        ("Lowest funded cash balance", lambda s: rands(needs[s]["min_closing_cash"])),
        ("Year 5 closing cash", lambda s: rands(results[s].annual_last("closing_cash")[4])),
    ]
    for label, fn in rows:
        w(f"| {label} | " + " | ".join(fn(s) for s in ("bootstrap", "bear", "base", "bull")) + " |")
    w("")

    w("## 5. Capital requirement\n")
    for m, label, amt in A.FUNDING_ROUNDS:
        w(f"- **Month {m} — {label}: {rands(amt)}**")
    w("")
    w(f"The base case's *unfunded* peak cash deficit is {rands(needs['base']['peak_deficit'])} "
      f"in month {needs['base']['peak_month']}. With the two rounds above, the lowest cash balance "
      f"is {rands(needs['base']['min_closing_cash'])} in month {needs['base']['min_closing_month']} — "
      "which is the single tightest point in the plan and the deadline for closing the growth round.")
    w("")
    w(f"The bear case needs {rands(needs['bear']['peak_deficit'])} and never comfortably repays it "
      "inside five years. That is the honest downside: if venue sell-through lands at half the plan, "
      "this is a business that survives but does not earn a venture return, and the right response "
      "is to hold headcount flat and run it as a niche brand rather than raise the growth round.")
    return "\n".join(L)


if __name__ == "__main__":
    path, results, needs = build_workbook()
    md = build_markdown(results, needs)
    md_path = os.path.join(OUT, "model-summary.md")
    with open(md_path, "w") as f:
        f.write(md + "\n")
    print("wrote", os.path.relpath(path))
    print("wrote", os.path.relpath(md_path))
